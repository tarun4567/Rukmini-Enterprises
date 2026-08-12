from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.db.models import F, Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from functools import wraps
import datetime
import json
import io

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from django.db import transaction
from .models import Category, Product, Bill, BillItem, StockHistory, Expense, ProductBatch, BatchSale, VendorPaymentHistory
from .forms import StockForm

# 1. Custom Role-Based Decorators
def admin_required(view_func):
    """Decorator for views that checks that the logged-in user is staff or superuser."""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.is_staff or request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        # Standard user is redirected to their day billing home
        messages.error(request, "Access Denied: You do not have administrator permissions.")
        return redirect('billing')
    return _wrapped_view


def standard_user_required(view_func):
    """Decorator for views that checks that the logged-in user is NOT staff or superuser (Billing/Counter staff)."""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not (request.user.is_staff or request.user.is_superuser):
            return view_func(request, *args, **kwargs)
        # Admins are redirected to their dashboard home
        messages.error(request, "Access Denied: Billing and Stocks screens are for standard billing operators only.")
        return redirect('dashboard')
    return _wrapped_view


# 2. Authentication Views
def login_view(request):
    if request.user.is_authenticated:
        if request.user.username == 'gst':
            return redirect('billing_records')
        if request.user.is_staff or request.user.is_superuser:
            return redirect('dashboard')
        return redirect('billing')
        
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {username}!")
                if user.username == 'gst':
                    return redirect('billing_records')
                elif user.is_staff or user.is_superuser:
                    return redirect('dashboard')
                return redirect('billing')
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect('login')


# 2b. Dashboard View (For Standard User Only)
@standard_user_required
def user_dashboard_view(request):
    daily_data, start_date, end_date, total_sales, total_due, total_expenditure = _get_user_dashboard_data(request)

    chart_labels = [d['display_date'] for d in daily_data]
    chart_sales = [d['sales'] for d in daily_data]
    chart_due = [d['due'] for d in daily_data]
    chart_exp = [d['expenditure'] for d in daily_data]

    # Calculate remaining balance: total sales minus expenses
    remaining_balance = total_sales - total_expenditure

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_sales': total_sales,
        'total_due': total_due,
        'total_expenditure': total_expenditure,
        'remaining_balance': remaining_balance,
        'daily_data': list(reversed(daily_data)), # Show most recent first in table
        'chart_labels': json.dumps(chart_labels),
        'chart_sales': json.dumps(chart_sales),
        'chart_due': json.dumps(chart_due),
        'chart_exp': json.dumps(chart_exp),
        'is_admin': False,
    }
    return render(request, 'standard_user_dashboard.html', context)


# 2c. Expenses View (For Standard User Only)
@standard_user_required
def user_expenses_view(request):
    if request.method == 'POST':
        # 1. Handle Deletion
        delete_id = request.POST.get('delete_id')
        if delete_id:
            expense_to_delete = get_object_or_404(Expense, pk=delete_id, recorded_by=request.user)
            comp = expense_to_delete.company_name
            amt = expense_to_delete.amount
            expense_to_delete.delete()
            messages.warning(request, f"Expense of \u20b9{amt:.2f} to '{comp}' was deleted successfully.")
            return redirect('user_expenses')
            
        # 2. Handle Creation
        company_name = request.POST.get('company_name', '').strip()
        amount_str   = request.POST.get('amount')
        date_paid_str = request.POST.get('date_paid')
        description  = request.POST.get('description', '').strip()
        
        if company_name and amount_str:
            try:
                amount = float(amount_str)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                else:
                    date_paid = timezone.localdate()
                    if date_paid_str:
                        date_paid = datetime.datetime.strptime(date_paid_str, '%Y-%m-%d').date()
                        
                    Expense.objects.create(
                        company_name=company_name,
                        amount=amount,
                        description=description,
                        date_paid=date_paid,
                        recorded_by=request.user
                    )
                    messages.success(request, f"Expense of \u20b9{amount:.2f} paid to '{company_name}' recorded successfully.")
                    return redirect('user_expenses')
            except ValueError:
                messages.error(request, "Invalid amount or date format.")
        else:
            messages.error(request, "Please fill in all required fields.")
            
    company_query = request.GET.get('company', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.localdate()
    start_date = today.replace(day=1)
    end_date = today
    
    try:
        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        pass
        
    try:
        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        pass

    expenses = Expense.objects.filter(
        date_paid__range=(start_date, end_date),
        recorded_by__is_staff=False,
        recorded_by__is_superuser=False
    )
    
    if company_query:
        expenses = expenses.filter(company_name__icontains=company_query)
        
    expenses = expenses.order_by('-date_paid', '-created_at')
    total_expense = float(expenses.aggregate(total=Sum('amount'))['total'] or 0.00)

    context = {
        'expenses': expenses,
        'company_query': company_query,
        'start_date': start_date,
        'end_date': end_date,
        'total_expense': total_expense,
        'today': today,
    }
    return render(request, 'user_expenses.html', context)


# 2d. Single Bill PDF Invoice Generation (For All Authenticated Users)
@login_required
def single_bill_pdf(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    
    buffer = io.BytesIO()
    # Tighten margins to fit more content cleanly on A4
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=1.2*cm, 
        leftMargin=1.2*cm,
        topMargin=1.2*cm, 
        bottomMargin=1.2*cm
    )
    elements = []

    # Get style sheet
    styles = getSampleStyleSheet()

    # Style definitions
    company_style = ParagraphStyle(
        'compname',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=3
    )
    company_sub = ParagraphStyle(
        'compsub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor('#475569')
    )
    invoice_title_style = ParagraphStyle(
        'invtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E293B'),
        alignment=TA_RIGHT
    )
    invoice_sub_style = ParagraphStyle(
        'invsub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor('#475569'),
        alignment=TA_RIGHT
    )
    section_heading = ParagraphStyle(
        'sechead',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=4
    )
    section_body = ParagraphStyle(
        'secbody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor('#334155')
    )
    tbl_header = ParagraphStyle(
        'tblhead',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.white
    )
    tbl_header_center = ParagraphStyle(
        'tblheadcenter',
        parent=tbl_header,
        alignment=TA_CENTER
    )
    tbl_header_right = ParagraphStyle(
        'tblheadright',
        parent=tbl_header,
        alignment=TA_RIGHT
    )
    tbl_cell = ParagraphStyle(
        'tblcell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#0F172A')
    )
    tbl_cell_center = ParagraphStyle(
        'tblcellcenter',
        parent=tbl_cell,
        alignment=TA_CENTER
    )
    tbl_cell_right = ParagraphStyle(
        'tblcellright',
        parent=tbl_cell,
        alignment=TA_RIGHT
    )
    totals_label_style = ParagraphStyle(
        'totalslabel',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor('#475569'),
        alignment=TA_RIGHT
    )
    totals_val_style = ParagraphStyle(
        'totalsval',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor('#0F172A'),
        alignment=TA_RIGHT
    )

    # 1. Header Layout (Side-by-side Table)
    local_created_at = timezone.localtime(bill.created_at)
    date_str = local_created_at.strftime('%d %b %Y, %I:%M %p')
    payment_mode_str = "Online / Card" if bill.payment_mode == 'ONLINE' else "Cash"
    
    header_left = [
        Paragraph("RUKMINI AGRO SEEDS", company_style),
        Paragraph("Chowdeshwari Nagara Near Nandini Bar Tekal Road Kolar - 563101<br/>"
                  "GSTIN:  29FAGPR9233F1Z8<br/>"
                  "Email: rukminiagroseeds2026@gmail.com | Phone: +91 75573 82143 , 9741900947", company_sub)
    ]
    
    header_right = [
        Paragraph("TAX INVOICE", invoice_title_style),
        Paragraph(f"<b>Invoice ID:</b> #{bill.id}<br/>"
                  f"<b>Date & Time:</b> {date_str}<br/>"
                  f"<b>Payment Mode:</b> {payment_mode_str}", invoice_sub_style)
    ]
    
    header_table = Table([[header_left, header_right]], colWidths=[11.0*cm, 7.6*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(header_table)

    # 2. Horizontal Divider Line
    divider = Table([['']], colWidths=[18.6*cm])
    divider.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,-1), 1.0, colors.HexColor('#CBD5E1')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(divider)
    elements.append(Spacer(1, 0.4*cm))

    # 3. Bill To Section
    bill_to_data = [
        [
            Paragraph("<b>BILL TO:</b>", section_heading),
            Paragraph("", section_body)
        ],
        [
            Paragraph(f"<b>Customer Name:</b> {bill.customer_name}", section_body),
            Paragraph(f"<b>Phone:</b> {bill.customer_phone or '\u2014'}", section_body)
        ],
        [
            Paragraph(f"<b>Address:</b> {bill.customer_address or '\u2014'}", section_body),
            Paragraph("", section_body)
        ]
    ]
    bill_to_table = Table(bill_to_data, colWidths=[9.3*cm, 9.3*cm])
    bill_to_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(bill_to_table)
    elements.append(Spacer(1, 0.5*cm))

    # 4. Items Table
    table_data = [[
        Paragraph("Sr.", tbl_header_center),
        Paragraph("Product / Item Description", tbl_header),
        Paragraph("Rate", tbl_header_right),
        Paragraph("Qty", tbl_header_center),
        Paragraph("Subtotal", tbl_header_right),
    ]]
    
    for i, item in enumerate(bill.items.select_related('product').all(), 1):
        prod_name = item.product.name if item.product else "Deleted Product"
        if item.product and item.product.size:
            prod_name += f" ({item.product.size})"
            
        table_data.append([
            Paragraph(str(i), tbl_cell_center),
            Paragraph(prod_name, tbl_cell),
            Paragraph(f"\u20b9{item.rate:.2f}", tbl_cell_right),
            Paragraph(str(item.quantity), tbl_cell_center),
            Paragraph(f"\u20b9{item.total:.2f}", tbl_cell_right),
        ])
        
    col_widths = [1.2*cm, 9.4*cm, 2.8*cm, 1.8*cm, 3.4*cm]  # Total width: 18.6cm
    items_table = Table(table_data, colWidths=col_widths)
    
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')), # Dark slate header
        ('ALIGN', (0,0), (-1,0), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 1.5, colors.HexColor('#0F172A')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')), 
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
    ]
    
    # Alternate row colors
    for r in range(1, len(table_data)):
        if r % 2 == 0:
            t_style.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#F8FAFC')))
            
    items_table.setStyle(TableStyle(t_style))
    elements.append(items_table)
    elements.append(Spacer(1, 0.4*cm))

    # 5. Totals Block (Structured right-aligned)
    if request.user.is_staff or request.user.is_superuser:
        totals_data = [
            [Paragraph("Total Bill Amount:", totals_label_style), Paragraph(f"\u20b9{bill.grand_total:.2f}", totals_val_style)],
            [Paragraph("GST (18%):", totals_label_style), Paragraph(f"\u20b9{bill.gst_amount:.2f}", totals_val_style)],
            [Paragraph("Total + GST:", totals_label_style), Paragraph(f"\u20b9{bill.grand_total_with_gst:.2f}", totals_val_style)],
            [Paragraph("Amount Paid:", totals_label_style), Paragraph(f"\u20b9{bill.grand_total_with_gst:.2f}", totals_val_style)],
        ]
    else:
        due_label = "Change Returned" if bill.amount_to_be_given >= 0 else "Balance Due"
        totals_data = [
            [Paragraph("Grand Total:", totals_label_style), Paragraph(f"\u20b9{bill.grand_total:.2f}", totals_val_style)],
            [Paragraph("Amount Paid:", totals_label_style), Paragraph(f"\u20b9{bill.amount_given:.2f}", totals_val_style)],
            [Paragraph(f"<b>{due_label}:</b>", totals_label_style), Paragraph(f"\u20b9{bill.abs_amount_to_be_given:.2f}", totals_val_style)],
        ]
        
    totals_col_widths = [4.2*cm, 2.8*cm]
    totals_sub_table = Table(totals_data, colWidths=totals_col_widths)
    totals_sub_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#F1F5F9')),
    ]))
    
    # Wrap totals in a container with a blank left cell to right-align it
    totals_container = Table([['', totals_sub_table]], colWidths=[11.6*cm, 7.0*cm])
    totals_container.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(totals_container)
    elements.append(Spacer(1, 0.8*cm))

    # 6. Terms and Signature Section
    terms_heading_style = ParagraphStyle(
        'termsh',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=2
    )
    terms_body_style = ParagraphStyle(
        'termsb',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor('#64748B')
    )
    
    terms_content = [
        Paragraph("Terms & Conditions:", terms_heading_style),
        Paragraph("1. Goods once sold will not be accepted for return or exchange.<br/>"
                  "2. Interest @ 18% p.a. will be charged if payment is not made on due date.<br/>"
                  "3. All disputes are subject to local court jurisdiction.", terms_body_style)
    ]
    
    signature_style = ParagraphStyle(
        'sigtext',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#475569'),
        alignment=TA_CENTER
    )
    
    signature_content = [
        Spacer(1, 0.8*cm),
        Paragraph("___________________________", signature_style),
        Spacer(1, 0.1*cm),
        Paragraph("Authorized Signatory", signature_style)
    ]
    
    footer_table = Table([[terms_content, signature_content]], colWidths=[11.6*cm, 7.0*cm])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(footer_table)
    
    thank_you_style = ParagraphStyle(
        'thanks', 
        fontSize=9, 
        fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#94A3B8'), 
        alignment=TA_CENTER,
        spaceBefore=15
    )
    elements.append(Paragraph("Thank you for your business!", thank_you_style))

    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bill_{bill.id}.pdf"'
    return response


# 3. Read-Only Stocks View (For Standard User Only)
@standard_user_required
def stocks_view(request):
    products = Product.objects.all().prefetch_related('batches')
    categories = Category.objects.all()
    active_category = request.GET.get('category')
    
    if active_category:
        products = products.filter(category__slug=active_category)
        
    context = {
        'products': products,
        'categories': categories,
        'active_category': active_category,
    }
    return render(request, 'stocks.html', context)


# 4. Day Billing View (For Standard User Only)
@standard_user_required
def billing_view(request):
    products = Product.objects.all()

    # Build active batches list for counter search
    active_batches_data = []
    for batch in ProductBatch.objects.filter(current_qty__gt=0).select_related('product'):
        active_batches_data.append({
            'id': batch.product.id,
            'batch_id': batch.id,
            'name': batch.product.name,
            'company': batch.product.company_name or '',
            'price': str(batch.selling_price),
            'stock': batch.current_qty,
            'size': batch.product.size or '',
            'batch_date': batch.stock_entered.strftime('%d %b %Y')
        })
    active_batches_json = json.dumps(active_batches_data)

    if request.method == 'POST':
        cust_name    = request.POST.get('customer_name', '').strip()
        cust_phone   = request.POST.get('customer_phone', '').strip()
        cust_address = request.POST.get('customer_address', '').strip()
        amount_given = float(request.POST.get('amount_given', 0.00))
        payment_mode = request.POST.get('payment_mode', 'CASH').strip()

        # Validate phone number if provided
        if cust_phone:
            if not cust_phone.isdigit() or len(cust_phone) != 10:
                messages.error(request, "Phone number must be exactly 10 digits (numbers only).")
                return render(request, 'billing.html', {'products': products, 'active_batches_json': active_batches_json})

        # Collect all item rows sent from the form
        batch_ids  = request.POST.getlist('batch_id[]')
        quantities = request.POST.getlist('quantity[]')
        rates      = request.POST.getlist('rate[]')

        if not batch_ids:
            messages.error(request, "Please add at least one product to the bill.")
            return render(request, 'billing.html', {'products': products, 'active_batches_json': active_batches_json})

        # Validate all items first before saving anything
        items_data = []
        grand_total = 0.0
        errors = []

        for i, bid in enumerate(batch_ids):
            try:
                qty = int(quantities[i])
            except (ValueError, IndexError):
                errors.append(f"Row {i+1}: Invalid quantity.")
                continue

            try:
                custom_rate = float(rates[i])
            except (ValueError, IndexError):
                errors.append(f"Row {i+1}: Invalid rate.")
                continue

            if qty <= 0:
                errors.append(f"Row {i+1}: Quantity must be > 0.")
                continue

            try:
                batch = ProductBatch.objects.get(id=bid)
            except ProductBatch.DoesNotExist:
                errors.append(f"Row {i+1}: Stock batch not found.")
                continue

            if batch.current_qty < qty:
                errors.append(
                    f"'{batch.product.name} (Batch: {batch.stock_entered})': Requested {qty} but only {batch.current_qty} available in this batch."
                )
                continue

            item_total = qty * custom_rate
            grand_total += float(item_total)
            items_data.append({
                'batch': batch,
                'qty': qty,
                'rate': custom_rate,
                'total': item_total
            })

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'billing.html', {'products': products, 'active_batches_json': active_batches_json})

        # All valid \u2014 create bill header
        amount_to_be_given = amount_given - grand_total
        bill = Bill.objects.create(
            customer_name=cust_name,
            customer_phone=cust_phone or None,
            customer_address=cust_address,
            grand_total=grand_total,
            amount_given=amount_given,
            amount_to_be_given=amount_to_be_given,
            payment_mode=payment_mode,
        )

        # Create line items and deduct stock
        with transaction.atomic():
            for item in items_data:
                batch = item['batch']
                bill_item = BillItem.objects.create(
                    bill=bill,
                    product=batch.product,
                    quantity=item['qty'],
                    rate=item['rate'],
                    total=item['total'],
                )
                
                BatchSale.objects.create(
                    bill_item=bill_item,
                    batch=batch,
                    quantity_sold=item['qty'],
                    purchase_rate=batch.purchase_rate
                )
                
                batch.current_qty -= item['qty']
                batch.save()
                
                # Sync product's stock_quantity
                batch.product.stock_quantity = max(0, batch.product.stock_quantity - item['qty'])
                batch.product.save()

        change_text = (
            f"Change: \u20b9{amount_to_be_given:.2f}"
            if amount_to_be_given >= 0
            else f"Due: \u20b9{abs(amount_to_be_given):.2f}"
        )
        messages.success(
            request,
            f"Bill #{bill.id} for {cust_name} created with {len(items_data)} item(s). {change_text}"
        )
        return redirect('billing_records')

    return render(request, 'billing.html', {'products': products, 'active_batches_json': active_batches_json})


# 4b. Day Billing Records (For Admin & Standard Users)
@login_required
def billing_records_view(request):
    date_str = request.GET.get('date')
    query_date = timezone.localdate()
    
    if date_str:
        try:
            query_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
            
    # Define start and end of the day in local timezone
    start_of_day = timezone.make_aware(datetime.datetime.combine(query_date, datetime.time.min))
    end_of_day = timezone.make_aware(datetime.datetime.combine(query_date, datetime.time.max))
    
    daily_bills = Bill.objects.filter(created_at__range=(start_of_day, end_of_day)).order_by('-created_at')
    
    # Calculate stats for admin only
    total_sales_today = 0.0
    total_received_today = 0.0
    total_dues_today = 0.0
    
    if request.user.is_staff or request.user.is_superuser:
        from django.db.models import Sum
        total_sales_today = float(daily_bills.aggregate(total=Sum('grand_total'))['total'] or 0.00) * 1.18
        total_received_today = float(daily_bills.aggregate(total=Sum('amount_given'))['total'] or 0.00) * 1.18
        total_dues_today = abs(float(daily_bills.filter(amount_to_be_given__lt=0).aggregate(total=Sum('amount_to_be_given'))['total'] or 0.00)) * 1.18
        
    context = {
        'daily_bills': daily_bills,
        'today': query_date,
        'total_sales_today': total_sales_today,
        'total_received_today': total_received_today,
        'total_dues_today': total_dues_today,
    }
    return render(request, 'billing_records.html', context)



# Helper to fetch user dashboard statistics and daily breakdown
def _get_user_dashboard_data(request):
    today = timezone.localdate()
    start_date_default = today.replace(day=1)
    end_date_default = today

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else start_date_default
    except ValueError:
        start_date = start_date_default

    try:
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else end_date_default
    except ValueError:
        end_date = end_date_default

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Limit to 90 days to prevent rendering lag
    if (end_date - start_date).days > 90:
        start_date = end_date - datetime.timedelta(days=90)

    daily_data = []
    current_day = start_date
    total_sales = 0.0
    total_due = 0.0
    total_expenditure = 0.0

    while current_day <= end_date:
        day_start = timezone.make_aware(datetime.datetime.combine(current_day, datetime.time.min))
        day_end = timezone.make_aware(datetime.datetime.combine(current_day, datetime.time.max))
        
        # Sales
        day_bills = Bill.objects.filter(created_at__range=(day_start, day_end))
        day_sales = float(day_bills.aggregate(total=Sum('grand_total'))['total'] or 0.00)
        
        # Due
        day_due = abs(float(day_bills.filter(amount_to_be_given__lt=0).aggregate(total=Sum('amount_to_be_given'))['total'] or 0.00))
        
        # Expenditure
        if request.user.is_staff or request.user.is_superuser:
            day_hist = StockHistory.objects.filter(date_entered=current_day, qty_added__gt=0)
            day_hist_total = sum(float(h.qty_added * h.product.purchase_price) for h in day_hist.select_related('product'))
            
            day_init = Product.objects.filter(stock_entered=current_day)
            day_init_total = sum(float(p.initial_quantity * p.purchase_price) for p in day_init)
        else:
            day_hist_total = 0.0
            day_init_total = 0.0
        
        if request.user.is_staff or request.user.is_superuser:
            day_manual_exp = float(Expense.objects.filter(date_paid=current_day).aggregate(total=Sum('amount'))['total'] or 0.00)
        else:
            day_manual_exp = float(Expense.objects.filter(
                date_paid=current_day,
                recorded_by__is_staff=False,
                recorded_by__is_superuser=False
            ).aggregate(total=Sum('amount'))['total'] or 0.00)
        
        day_exp = day_hist_total + day_init_total + day_manual_exp
        
        total_sales += day_sales
        total_due += day_due
        total_expenditure += day_exp
        
        daily_data.append({
            'date': current_day,
            'display_date': current_day.strftime('%b %d, %Y'),
            'sales': day_sales,
            'due': day_due,
            'expenditure': day_exp,
            'net': day_sales - day_exp
        })
        
        current_day += datetime.timedelta(days=1)

    return daily_data, start_date, end_date, total_sales, total_due, total_expenditure



# 4bc. User-Side Dashboard Excel Export (For All Authenticated Users)
@login_required
def user_dashboard_excel(request):
    daily_data, start_date, end_date, total_sales, total_due, total_expenditure = _get_user_dashboard_data(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts Report"

    # Styles
    header_fill = PatternFill("solid", fgColor="2563EB") # Royal Blue
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_font = Font(bold=True, size=11)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Rukmini Enterprises \u2014 Accounts Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center

    ws.append([]) # empty row

    # Summary Metrics Table
    ws.append(["Summary Metrics", "", "", "", ""])
    ws.merge_cells("A4:E4")
    ws["A4"].font = Font(bold=True, size=12)
    ws["A4"].fill = PatternFill("solid", fgColor="F3F4F6")

    ws.append(["Total Sales Done", "Total Outstanding Dues", "Total Expenditure", "Net Balance", ""])
    ws.merge_cells("D5:E5")
    for col in range(1, 6):
        ws.cell(row=5, column=col).font = bold
        ws.cell(row=5, column=col).border = border

    ws.append([total_sales, total_due, total_expenditure, total_sales - total_expenditure, ""])
    ws.merge_cells("D6:E6")
    for col in range(1, 6):
        cell = ws.cell(row=6, column=col)
        cell.font = summary_font
        cell.border = border
        if col == 4:
            if (total_sales - total_expenditure) >= 0:
                cell.font = Font(bold=True, color="15803D") # green
            else:
                cell.font = Font(bold=True, color="B91C1C") # red

    ws.append([]) # empty row

    # Headers
    headers = ["Date", "Sales Done (INR)", "Outstanding Dues (INR)", "Expenditure (INR)", "Net Balance (INR)"]
    ws.append(headers)
    for col, cell in enumerate(ws[8], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for d in daily_data:
        ws.append([
            d['display_date'],
            d['sales'],
            d['due'],
            d['expenditure'],
            d['net']
        ])
        row = ws.max_row
        for col_idx, cell in enumerate(ws[row], 1):
            cell.border = border
            if col_idx > 1:
                cell.alignment = right
                cell.number_format = '\u20b9#,##0.00'
            else:
                cell.alignment = center

    # Column widths
    col_widths = [16, 20, 24, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"accounts_report_{start_date}_to_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# 4bd. User-Side Dashboard PDF Export (For All Authenticated Users)
@login_required
def user_dashboard_pdf(request):
    daily_data, start_date, end_date, total_sales, total_due, total_expenditure = _get_user_dashboard_data(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = []

    title_style = ParagraphStyle('title', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=3)
    sub_style   = ParagraphStyle('sub',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=12)
    label_style = ParagraphStyle('lbl',   fontSize=9,  fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#475569'), alignment=TA_CENTER)
    val_style   = ParagraphStyle('val',   fontSize=12, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#0F172A'), alignment=TA_CENTER)

    elements.append(Paragraph("Rukmini Enterprises", title_style))
    elements.append(Paragraph(f"Monthly Accounts Report (Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')})", sub_style))

    # Summary table
    summary_data = [
        [
            Paragraph("Total Sales Done", label_style),
            Paragraph("Total Outstanding Dues", label_style),
            Paragraph("Total Expenditure", label_style),
            Paragraph("Net Balance", label_style)
        ],
        [
            Paragraph(f"\u20b9{total_sales:,.2f}", val_style),
            Paragraph(f"\u20b9{total_due:,.2f}", val_style),
            Paragraph(f"\u20b9{total_expenditure:,.2f}", val_style),
            Paragraph(f"\u20b9{total_sales - total_expenditure:,.2f}", ParagraphStyle('net_val', parent=val_style, textColor=colors.HexColor('#15803D') if (total_sales - total_expenditure) >= 0 else colors.HexColor('#B91C1C')))
        ]
    ]

    sum_table = Table(summary_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 15))

    # Daily Ledger details
    th_style = ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER)
    td_style = ParagraphStyle('td', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#334155'))
    td_num_style = ParagraphStyle('td_num', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#334155'), alignment=TA_RIGHT)

    table_data = [[
        Paragraph("Date", th_style),
        Paragraph("Sales Done", th_style),
        Paragraph("Outstanding Dues", th_style),
        Paragraph("Expenditure", th_style),
        Paragraph("Net Balance", th_style),
    ]]

    for d in daily_data:
        table_data.append([
            Paragraph(d['display_date'], td_style),
            Paragraph(f"\u20b9{d['sales']:,.2f}", td_num_style),
            Paragraph(f"\u20b9{d['due']:,.2f}", td_num_style),
            Paragraph(f"\u20b9{d['expenditure']:,.2f}", td_num_style),
            Paragraph(f"\u20b9{d['net']:,.2f}", ParagraphStyle('net_td', parent=td_num_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#16A34A') if d['net'] >= 0 else colors.HexColor('#DC2626'))),
        ])

    ledger_table = Table(table_data, colWidths=[3.5*cm, 3.6*cm, 3.7*cm, 3.6*cm, 3.6*cm])
    ledger_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563EB')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
    ]))
    elements.append(ledger_table)

    doc.build(elements)
    pdf_val = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    filename = f"accounts_report_{start_date}_to_{end_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf_val)
    return response


# 4c. Clear Due View (For Standard User Only)
@standard_user_required
def clear_due_view(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    if request.method == 'POST':
        clear_amount_str = request.POST.get('clear_amount')
        if clear_amount_str:
            try:
                clear_amount = float(clear_amount_str)
                if clear_amount > 0:
                    max_due = abs(float(bill.amount_to_be_given))
                    # Prevent floating point inaccuracies and overpaying
                    if clear_amount > round(max_due, 2):
                        clear_amount = max_due
                    
                    bill.amount_given = float(bill.amount_given) + clear_amount
                    bill.amount_to_be_given = float(bill.amount_given) - float(bill.total)
                    bill.save()
                    
                    if bill.amount_to_be_given >= 0:
                        messages.success(request, f"Payment of \u20b9{clear_amount:.2f} received. Bill #{bill.id} is now fully cleared!")
                    else:
                        messages.success(request, f"Payment of \u20b9{clear_amount:.2f} received for Bill #{bill.id}. Remaining due: \u20b9{abs(bill.amount_to_be_given):.2f}")
                else:
                    messages.error(request, "Payment amount must be greater than zero.")
            except ValueError:
                messages.error(request, "Invalid payment amount format.")
        else:
            messages.error(request, "No payment amount was specified.")
            
        local_created_at = timezone.localtime(bill.created_at)
        bill_date_str = local_created_at.strftime('%Y-%m-%d')
        return redirect(f"/billing/records/?date={bill_date_str}")
    return redirect('billing_records')


# 5. Dashboard View (Admin Only)
@admin_required
def dashboard_view(request):
    daily_data, start_date, end_date, total_sales, total_due, total_expenditure = _get_user_dashboard_data(request)

    chart_labels = [d['display_date'] for d in daily_data]
    chart_sales = [d['sales'] for d in daily_data]
    chart_due = [d['due'] for d in daily_data]
    chart_exp = [d['expenditure'] for d in daily_data]

    # Calculate Cash Sales and Credit Sales (Outstanding Due)
    credit_sale = total_due
    cash_sale   = max(0.00, total_sales - total_due)
    total_sales_combined = cash_sale + credit_sale

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'cash_sale': cash_sale,
        'credit_sale': credit_sale,
        'total_sales': total_sales_combined,
        'total_due': total_due,
        'total_expenditure': total_expenditure,
        'daily_data': list(reversed(daily_data)), # Show most recent first in table
        'chart_labels': json.dumps(chart_labels),
        'chart_sales': json.dumps(chart_sales),
        'chart_due': json.dumps(chart_due),
        'chart_exp': json.dumps(chart_exp),
        'is_admin': True,
    }
    return render(request, 'user_dashboard.html', context)


# 5c. Admin Expenses List/Manage View (Admin Only)
@admin_required
def admin_expenses_view(request):
    # Handle POST requests (Create or Delete)
    if request.method == 'POST':
        # 1. Handle Deletion
        delete_id = request.POST.get('delete_id')
        if delete_id:
            expense_to_delete = get_object_or_404(Expense, pk=delete_id)
            comp = expense_to_delete.company_name
            amt = expense_to_delete.amount
            expense_to_delete.delete()
            messages.warning(request, f"Expense of \u20b9{amt:.2f} to '{comp}' was deleted successfully.")
            return redirect('admin_expenses')
            
        # 2. Handle Creation
        company_name = request.POST.get('company_name', '').strip()
        amount_str   = request.POST.get('amount')
        date_paid_str = request.POST.get('date_paid')
        description  = request.POST.get('description', '').strip()
        
        if company_name and amount_str:
            try:
                amount = float(amount_str)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                else:
                    date_paid = timezone.localdate()
                    if date_paid_str:
                        date_paid = datetime.datetime.strptime(date_paid_str, '%Y-%m-%d').date()
                        
                    Expense.objects.create(
                         company_name=company_name,
                         amount=amount,
                         description=description,
                         date_paid=date_paid,
                         recorded_by=request.user
                    )
                    messages.success(request, f"Expense of \u20b9{amount:.2f} paid to '{company_name}' recorded successfully.")
                    return redirect('admin_expenses')
            except ValueError:
                messages.error(request, "Invalid amount or date format.")
        else:
            messages.error(request, "Please fill in all required fields.")
            
    # Search and date filters
    company_query = request.GET.get('company', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.localdate()
    start_date = today.replace(day=1)
    end_date = today
    
    try:
        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        pass
        
    try:
        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        pass

    expenses = Expense.objects.filter(
        date_paid__range=(start_date, end_date),
        recorded_by__is_staff=True
    )
    
    if company_query:
        expenses = expenses.filter(company_name__icontains=company_query)
        
    expenses = expenses.order_by('-date_paid', '-created_at')
    
    # Calculate sum
    total_expense = float(expenses.aggregate(total=Sum('amount'))['total'] or 0.00)

    context = {
        'expenses': expenses,
        'company_query': company_query,
        'start_date': start_date,
        'end_date': end_date,
        'total_expense': total_expense,
        'today': today,
    }
    return render(request, 'admin_expenses.html', context)


# -- Vendor Manager View (Admin Only) -------------------------
@admin_required
def vendor_view(request):
    today = timezone.localdate()
    three_days_later = today + datetime.timedelta(days=3)

    if request.method == 'POST':
        action = request.POST.get('action')
        company_name = request.POST.get('company_name')

        if action == 'update_due_date':
            due_date_str = request.POST.get('vendor_due_date')
            if company_name is not None and due_date_str:
                try:
                    due_date = datetime.datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    if company_name in [None, '', 'Unknown Vendor']:
                        products_to_update = Product.objects.filter(
                            Q(company_name__isnull=True) | Q(company_name=''),
                            remaining_amount_to_vendor__gt=0
                        )
                    else:
                        products_to_update = Product.objects.filter(
                            company_name=company_name,
                            remaining_amount_to_vendor__gt=0
                        )
                    
                    count = products_to_update.count()
                    for product in products_to_update:
                        product.vendor_due_date = due_date
                        product.save()
                    
                    display_name = company_name if company_name not in [None, '', 'Unknown Vendor'] else 'Unknown Vendor'
                    messages.success(request, f"Updated due date to {due_date.strftime('%d %b %Y')} for all {count} active items of vendor '{display_name}'.")
                    return redirect('admin_vendor')
                except ValueError:
                    messages.error(request, "Invalid date format.")
        else:
            pay_amount_str = request.POST.get('pay_amount')
            desc = request.POST.get('description', '').strip()
            if company_name is not None and pay_amount_str:
                try:
                    pay_amount = float(pay_amount_str)
                    
                    if company_name in [None, '', 'Unknown Vendor']:
                        products_to_pay = Product.objects.filter(
                            Q(company_name__isnull=True) | Q(company_name=''),
                            remaining_amount_to_vendor__gt=0
                        ).order_by('vendor_due_date', 'id')
                    else:
                        products_to_pay = Product.objects.filter(
                            company_name=company_name,
                            remaining_amount_to_vendor__gt=0
                        ).order_by('vendor_due_date', 'id')

                    total_remaining = float(sum(p.remaining_amount_to_vendor for p in products_to_pay))

                    if pay_amount <= 0:
                        messages.error(request, "Payment amount must be greater than zero.")
                    elif pay_amount > total_remaining:
                        messages.error(request, f"Payment amount exceeds the total remaining balance of \u20b9{total_remaining:.2f}.")
                    else:
                        remaining_payment = pay_amount
                        with transaction.atomic():
                            for product in products_to_pay:
                                if remaining_payment <= 0:
                                    break
                                current_remaining = float(product.remaining_amount_to_vendor or 0.00)
                                payment_to_apply = min(remaining_payment, current_remaining)
                                
                                product.remaining_amount_to_vendor = max(0.00, current_remaining - payment_to_apply)
                                product.amount_paid_to_vendor = float(product.amount_paid_to_vendor or 0.00) + payment_to_apply
                                product.save()

                                # Log in VendorPaymentHistory
                                VendorPaymentHistory.objects.create(
                                    product=product,
                                    amount_paid=payment_to_apply,
                                    payment_date=today,
                                    description=desc or None,
                                    recorded_by=request.user
                                )
                                remaining_payment -= payment_to_apply

                        display_name = company_name if company_name not in [None, '', 'Unknown Vendor'] else 'Unknown Vendor'
                        messages.success(request, f"Successfully paid \u20b9{pay_amount:.2f} to vendor '{display_name}'.")
                        return redirect('admin_vendor')
                except ValueError:
                    messages.error(request, "Invalid payment amount entered.")

    search_query = request.GET.get('search', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    outstanding_products = Product.objects.filter(
        remaining_amount_to_vendor__gt=0
    )
    if search_query:
        outstanding_products = outstanding_products.filter(
            Q(company_name__icontains=search_query) | Q(name__icontains=search_query)
        )
    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__gte=start_date)
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__lte=end_date)
        except ValueError:
            pass
    outstanding_products = outstanding_products.order_by('vendor_due_date', 'company_name')

    # Group outstanding products by company_name
    grouped_vendors = {}
    for p in outstanding_products:
        c_name = p.company_name or "Unknown Vendor"
        if c_name not in grouped_vendors:
            grouped_vendors[c_name] = {
                'company_name': c_name,
                'name': c_name,
                'products': [],
                'amount_paid_to_vendor': 0.0,
                'remaining_amount_to_vendor': 0.0,
                'vendor_due_date': None,
            }
        gv = grouped_vendors[c_name]
        gv['products'].append(p)
        gv['amount_paid_to_vendor'] += float(p.amount_paid_to_vendor or 0.00)
        gv['remaining_amount_to_vendor'] += float(p.remaining_amount_to_vendor or 0.00)
        if p.vendor_due_date:
            if gv['vendor_due_date'] is None or p.vendor_due_date < gv['vendor_due_date']:
                gv['vendor_due_date'] = p.vendor_due_date

    outstanding_vendors_list = list(grouped_vendors.values())
    outstanding_vendors_list.sort(
        key=lambda x: (
            x['vendor_due_date'] is None,
            x['vendor_due_date'] or datetime.date.max,
            x['company_name']
        )
    )

    # Repayments due within 3 days (warning pop-up trigger list)
    upcoming_products = Product.objects.filter(
        remaining_amount_to_vendor__gt=0,
        vendor_due_date__range=(today, three_days_later)
    ).order_by('vendor_due_date')

    upcoming_dues_grouped = {}
    for p in upcoming_products:
        c_name = p.company_name or "Unknown Vendor"
        if c_name not in upcoming_dues_grouped:
            upcoming_dues_grouped[c_name] = {
                'company_name': c_name,
                'name': c_name,
                'remaining_amount_to_vendor': 0.0,
                'vendor_due_date': p.vendor_due_date
            }
        upcoming_dues_grouped[c_name]['remaining_amount_to_vendor'] += float(p.remaining_amount_to_vendor or 0.00)
    upcoming_dues = list(upcoming_dues_grouped.values())

    # Also include already-overdue payments
    overdue_products = Product.objects.filter(
        remaining_amount_to_vendor__gt=0,
        vendor_due_date__lt=today
    ).order_by('vendor_due_date')

    overdue_dues_grouped = {}
    for p in overdue_products:
        c_name = p.company_name or "Unknown Vendor"
        if c_name not in overdue_dues_grouped:
            overdue_dues_grouped[c_name] = {
                'company_name': c_name,
                'name': c_name,
                'remaining_amount_to_vendor': 0.0,
                'vendor_due_date': p.vendor_due_date
            }
        overdue_dues_grouped[c_name]['remaining_amount_to_vendor'] += float(p.remaining_amount_to_vendor or 0.00)
    overdue_dues = list(overdue_dues_grouped.values())

    # Retrieve vendor repayment logs history
    filter_company = request.GET.get('filter_company', '').strip()
    payment_history = VendorPaymentHistory.objects.select_related('product').all()
    if filter_company:
        payment_history = payment_history.filter(product__company_name__icontains=filter_company)
    payment_history = payment_history.order_by('-created_at')[:50]

    # Calculate summary statistics for the dashboard cards
    total_remaining = float(Product.objects.filter(remaining_amount_to_vendor__gt=0).aggregate(total=Sum('remaining_amount_to_vendor'))['total'] or 0.00)
    total_paid_active = float(Product.objects.filter(remaining_amount_to_vendor__gt=0).aggregate(total=Sum('amount_paid_to_vendor'))['total'] or 0.00)
    total_paid_all = float(Product.objects.filter(amount_paid_to_vendor__gt=0).aggregate(total=Sum('amount_paid_to_vendor'))['total'] or 0.00)
    outstanding_count = Product.objects.filter(remaining_amount_to_vendor__gt=0).values('company_name').distinct().count()

    context = {
        'outstanding_vendors': outstanding_vendors_list,
        'upcoming_dues': upcoming_dues,
        'overdue_dues': overdue_dues,
        'payment_history': payment_history,
        'filter_company': filter_company,
        'search_query': search_query,
        'start_date_str': start_date_str,
        'end_date_str': end_date_str,
        'today': today,
        'total_remaining': total_remaining,
        'total_paid_active': total_paid_active,
        'total_paid_all': total_paid_all,
        'outstanding_count': outstanding_count,
    }
    return render(request, 'admin_vendor.html', context)


# -- Vendor Report: Excel Download (Admin Only) ---------------
@admin_required
def vendor_report_excel(request):
    search_query = request.GET.get('search', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    outstanding_products = Product.objects.filter(
        remaining_amount_to_vendor__gt=0
    )
    if search_query:
        outstanding_products = outstanding_products.filter(
            Q(company_name__icontains=search_query) | Q(name__icontains=search_query)
        )
    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__gte=start_date)
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__lte=end_date)
        except ValueError:
            pass
    outstanding_products = outstanding_products.order_by('vendor_due_date', 'company_name')

    # Group outstanding products by company_name
    grouped_vendors = {}
    for p in outstanding_products:
        c_name = p.company_name or "Unknown Vendor"
        if c_name not in grouped_vendors:
            grouped_vendors[c_name] = {
                'company_name': c_name,
                'product_names': [],
                'total_paid': 0.0,
                'total_remaining': 0.0,
                'earliest_due_date': None,
            }
        gv = grouped_vendors[c_name]
        gv['product_names'].append(p.name)
        gv['total_paid'] += float(p.amount_paid_to_vendor or 0.00)
        gv['total_remaining'] += float(p.remaining_amount_to_vendor or 0.00)
        if p.vendor_due_date:
            if gv['earliest_due_date'] is None or p.vendor_due_date < gv['earliest_due_date']:
                gv['earliest_due_date'] = p.vendor_due_date

    outstanding_vendors_list = list(grouped_vendors.values())
    outstanding_vendors_list.sort(
        key=lambda x: (
            x['earliest_due_date'] is None,
            x['earliest_due_date'] or datetime.date.max,
            x['company_name']
        )
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Credit Report"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    overdue_fill = PatternFill("solid", fgColor="FEE2E2")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "Rukmini Enterprises \u2014 Vendor Credit Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {timezone.localdate().strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center
    ws.append([])

    headers = ["#", "Vendor / Company", "Stock Item", "Paid (\u20b9)", "Remaining (\u20b9)", "Due Date"]
    ws.append(headers)
    for col, cell in enumerate(ws[4], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    today = timezone.localdate()
    for i, v in enumerate(outstanding_vendors_list, 1):
        dt_str = v['earliest_due_date'].strftime('%d %b %Y') if v['earliest_due_date'] else '\u2014'
        ws.append([
            i,
            v['company_name'],
            ", ".join(v['product_names']),
            v['total_paid'],
            v['total_remaining'],
            dt_str,
        ])
        row = ws.max_row
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if v['earliest_due_date'] and v['earliest_due_date'] < today:
            for cell in ws[row]:
                cell.fill = overdue_fill

    col_widths = [6, 25, 28, 16, 16, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="vendor_report_{timezone.localdate()}.xlsx"'
    wb.save(response)
    return response


# -- Vendor Report: PDF Download (Admin Only) -----------------
@admin_required
def vendor_report_pdf(request):
    search_query = request.GET.get('search', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    outstanding_products = Product.objects.filter(
        remaining_amount_to_vendor__gt=0
    )
    if search_query:
        outstanding_products = outstanding_products.filter(
            Q(company_name__icontains=search_query) | Q(name__icontains=search_query)
        )
    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__gte=start_date)
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            outstanding_products = outstanding_products.filter(vendor_due_date__lte=end_date)
        except ValueError:
            pass
    outstanding_products = outstanding_products.order_by('vendor_due_date', 'company_name')

    # Group outstanding products by company_name
    grouped_vendors = {}
    for p in outstanding_products:
        c_name = p.company_name or "Unknown Vendor"
        if c_name not in grouped_vendors:
            grouped_vendors[c_name] = {
                'company_name': c_name,
                'product_names': [],
                'total_paid': 0.0,
                'total_remaining': 0.0,
                'earliest_due_date': None,
            }
        gv = grouped_vendors[c_name]
        gv['product_names'].append(p.name)
        gv['total_paid'] += float(p.amount_paid_to_vendor or 0.00)
        gv['total_remaining'] += float(p.remaining_amount_to_vendor or 0.00)
        if p.vendor_due_date:
            if gv['earliest_due_date'] is None or p.vendor_due_date < gv['earliest_due_date']:
                gv['earliest_due_date'] = p.vendor_due_date

    outstanding_vendors_list = list(grouped_vendors.values())
    outstanding_vendors_list.sort(
        key=lambda x: (
            x['earliest_due_date'] is None,
            x['earliest_due_date'] or datetime.date.max,
            x['company_name']
        )
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []

    title_style = ParagraphStyle('vtitle', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=3)
    sub_style   = ParagraphStyle('vsub',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=12)

    elements.append(Paragraph("Rukmini Enterprises - Vendor Credit Report", title_style))
    elements.append(Paragraph(f"Generated: {timezone.localdate().strftime('%d %b %Y')}", sub_style))

    # Total width = 27.7cm
    col_widths_pdf = [1.2*cm, 7.0*cm, 8.0*cm, 4.0*cm, 4.5*cm, 3.0*cm]
    col_headers = ["#", "Vendor / Company", "Stock Item", "Paid", "Remaining Due", "Due Date"]
    data = [col_headers]

    t_styles = [
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  8),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
        ('VALIGN',        (0,0), (-1,0),  'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  6),
        ('BOTTOMPADDING', (0,0), (-1,0),  6),
    ]

    today = timezone.localdate()
    for i, v in enumerate(outstanding_vendors_list, 1):
        dt_str = v['earliest_due_date'].strftime('%d %b %Y') if v['earliest_due_date'] else '\u2014'
        row_idx = len(data)
        products_summary = ", ".join(v['product_names'])
        if len(products_summary) > 45:
            products_summary = products_summary[:42] + "..."
        data.append([
            str(i),
            v['company_name'],
            products_summary,
            f"Rs.{v['total_paid']:,.2f}",
            f"Rs.{v['total_remaining']:,.2f}",
            dt_str,
        ])

        if v['earliest_due_date'] and v['earliest_due_date'] < today:
            t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#FEE2E2')))
        else:
            bg_color = colors.white if i % 2 != 0 else colors.HexColor('#F8FAFC')
            t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))

        t_styles.extend([
            ('FONTNAME',      (0, row_idx), (-1, row_idx), 'Helvetica'),
            ('FONTSIZE',      (0, row_idx), (-1, row_idx), 7.5),
            ('ALIGN',         (0, row_idx), (0, row_idx),  'CENTER'),
            ('ALIGN',         (1, row_idx), (2, row_idx),  'LEFT'),
            ('ALIGN',         (3, row_idx), (4, row_idx),  'RIGHT'),
            ('ALIGN',         (5, row_idx), (5, row_idx),  'CENTER'),
            ('VALIGN',        (0, row_idx), (-1, row_idx), 'MIDDLE'),
            ('TOPPADDING',    (0, row_idx), (-1, row_idx), 5),
            ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 5),
        ])

    if len(data) == 1:
        data.append(["No outstanding vendor accounts found."] + [""] * 5)
        t_styles.append(('SPAN', (0, 1), (-1, 1)))

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    t_styles.extend([
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
    ])
    tbl.setStyle(TableStyle(t_styles))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="vendor_report_{timezone.localdate()}.pdf"'
    return response


# 6. Stock & Stock 1 Views (Admin Only)
@admin_required
def stock_add_view(request):
    categories = Category.objects.all()
    products = Product.objects.all().order_by('-id')
    
    if request.method == 'POST':
        form = StockForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                name = form.cleaned_data.get('name').strip()
                company_name = form.cleaned_data.get('company_name', '').strip()
                
                # Check for existing product with same name and company name (case-insensitive)
                existing_product = Product.objects.filter(
                    name__iexact=name,
                    company_name__iexact=company_name
                ).first()
                
                # Retrieve fields to determine new batch size and costs
                new_qty = form.cleaned_data.get('initial_quantity') or 0
                vendor_cost = form.cleaned_data.get('vendor_cost') or 0.00
                selling_price = form.cleaned_data.get('selling_price') or 0.00
                purchase_amount = float(new_qty) * float(vendor_cost)
                payment_made = float(form.cleaned_data.get('amount_paid_to_vendor') or 0.00)
                
                # Validation: Payment made cannot exceed current purchase total
                if payment_made > purchase_amount:
                    form.add_error('amount_paid_to_vendor', f"Payment made (\u20b9{payment_made}) cannot be greater than the purchase total (\u20b9{purchase_amount}).")
                    return render(request, 'stock.html', {'form': form, 'categories': categories, 'products': products})
                
                if existing_product:
                    # Update existing product inventory
                    existing_product.initial_quantity += new_qty
                    existing_product.stock_quantity += new_qty
                    existing_product.vendor_cost = vendor_cost
                    existing_product.selling_price = selling_price
                    
                    current_outstanding = purchase_amount - payment_made
                    
                    # Accumulate totals
                    existing_product.total_vendor_amount = float(existing_product.total_vendor_amount or 0.00) + purchase_amount
                    existing_product.amount_paid_to_vendor = float(existing_product.amount_paid_to_vendor or 0.00) + payment_made
                    existing_product.remaining_amount_to_vendor = float(existing_product.remaining_amount_to_vendor or 0.00) + current_outstanding
                    
                    # Save image/photo/description/size if updated in form
                    if form.cleaned_data.get('image'):
                        existing_product.image = form.cleaned_data.get('image')
                    if form.cleaned_data.get('photo'):
                        existing_product.photo = form.cleaned_data.get('photo')
                    if form.cleaned_data.get('size'):
                        existing_product.size = form.cleaned_data.get('size')
                    if form.cleaned_data.get('description'):
                        existing_product.description = form.cleaned_data.get('description')
                    if form.cleaned_data.get('category'):
                        existing_product.category = form.cleaned_data.get('category')
                        
                    existing_product.save()
                    
                    # Create the ProductBatch for this new stock addition
                    ProductBatch.objects.create(
                        product=existing_product,
                        purchase_rate=vendor_cost,
                        selling_price=selling_price,
                        initial_qty=new_qty,
                        current_qty=new_qty,
                        stock_entered=form.cleaned_data.get('stock_entered') or timezone.localdate()
                    )
                    
                    # Log stock history
                    StockHistory.objects.create(
                        product=existing_product,
                        date_entered=form.cleaned_data.get('stock_entered') or timezone.localdate(),
                        qty_added=new_qty,
                        qty_after=existing_product.stock_quantity,
                    )
                    
                    messages.success(request, f"Added {new_qty} units to existing stock item '{existing_product.name}'. Vendor balance updated.")
                else:
                    # Create completely new product
                    product = form.save(commit=False)
                    product.stock_quantity = product.initial_quantity
                    
                    product.total_vendor_amount = purchase_amount
                    product.remaining_amount_to_vendor = max(0.00, purchase_amount - payment_made)
                    product.save()
                    
                    ProductBatch.objects.create(
                        product=product,
                        purchase_rate=product.vendor_cost or 0.00,
                        selling_price=product.selling_price or 0.00,
                        initial_qty=product.initial_quantity,
                        current_qty=product.initial_quantity,
                        stock_entered=product.stock_entered or timezone.localdate()
                    )
                    
                    messages.success(request, f"Stock item '{product.name}' was added successfully with its first batch.")
            return redirect('stock_add')
    else:
        form = StockForm()
        
    # Fetch unique company names with their latest vendor details to support auto-fill
    existing_vendors_data = []
    seen_companies = set()
    for p in Product.objects.exclude(company_name__isnull=True).exclude(company_name="").order_by('-id'):
        cname = p.company_name.strip()
        cname_lower = cname.lower()
        if cname_lower not in seen_companies:
            seen_companies.add(cname_lower)
            existing_vendors_data.append({
                'company_name': cname,
                'due_date': p.vendor_due_date.strftime('%Y-%m-%d') if p.vendor_due_date else '',
                'total_vendor_amount': float(p.total_vendor_amount) if p.total_vendor_amount is not None else '',
                'amount_paid_to_vendor': float(p.amount_paid_to_vendor) if p.amount_paid_to_vendor is not None else '',
                'remaining_amount_to_vendor': float(p.remaining_amount_to_vendor) if p.remaining_amount_to_vendor is not None else '',
            })

    context = {
        'form': form,
        'products': products,
        'categories': categories,
        'existing_vendors_json': json.dumps(existing_vendors_data),
    }
    return render(request, 'stock.html', context)


@admin_required
def stock1_view(request):
    products = Product.objects.all().order_by('-id')
    return render(request, 'stock1.html', {'products': products})


@admin_required
def stock_edit_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.all()
    products = Product.objects.all().order_by('-id')
    
    if request.method == 'POST':
        old_initial_qty = product.initial_quantity
        form = StockForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            with transaction.atomic():
                # Extract initial values before POST changes are committed
                old_amount_paid = float(form.initial.get('amount_paid_to_vendor') or 0.00)
                old_remaining = float(form.initial.get('remaining_amount_to_vendor') or 0.00)
                old_total = float(form.initial.get('total_vendor_amount') or 0.00)

                product = form.save(commit=False)
                
                # Check if a new batch was added through the consignment fields
                add_qty_str = request.POST.get('add_batch_qty', '').strip()
                add_qty = int(add_qty_str) if add_qty_str and add_qty_str.isdigit() else 0
                
                if add_qty > 0:
                    add_cost_str = request.POST.get('add_batch_cost', '').strip()
                    try:
                        add_cost = float(add_cost_str) if add_cost_str else float(product.vendor_cost or 0.00)
                    except ValueError:
                        add_cost = float(product.vendor_cost or 0.00)
                        
                    # Calculate new consignment purchase amount
                    purchase_total = float(add_qty) * float(add_cost)
                    
                    # Read batch payment from POST parameters
                    add_paid_str = request.POST.get('add_batch_paid', '0.00').strip()
                    try:
                        add_paid = float(add_paid_str) if add_paid_str else 0.00
                    except ValueError:
                        add_paid = 0.00
                        
                    # Validate: Payment cannot be greater than batch total
                    if add_paid > purchase_total:
                        form.add_error(None, f"Payment made for this batch (\u20b9{add_paid}) cannot be greater than the batch total (\u20b9{purchase_total}).")
                        stock_history = product.stock_history.all()
                        batches = product.batches.all().order_by('-created_at')
                        return render(request, 'stock.html', {
                            'form': form,
                            'product': product,
                            'categories': categories,
                            'products': products,
                            'stock_history': stock_history,
                            'batches': batches,
                            'is_edit': True,
                        })
                        
                    add_sell_str = request.POST.get('add_batch_sell', '').strip()
                    try:
                        add_sell = float(add_sell_str) if add_sell_str else float(product.selling_price)
                    except ValueError:
                        add_sell = float(product.selling_price)
                        
                    add_date_str = request.POST.get('add_batch_date', '').strip()
                    add_date = timezone.localdate()
                    if add_date_str:
                        try:
                            add_date = datetime.datetime.strptime(add_date_str, "%Y-%m-%d").date()
                        except ValueError:
                            pass
                            
                    # Create the new batch
                    ProductBatch.objects.create(
                        product=product,
                        purchase_rate=add_cost,
                        selling_price=add_sell,
                        initial_qty=add_qty,
                        current_qty=add_qty,
                        stock_entered=add_date
                    )
                    
                    # Update product totals
                    product.initial_quantity += add_qty
                    product.stock_quantity += add_qty
                    product.vendor_cost = add_cost  # update to latest vendor cost
                    product.selling_price = add_sell  # update product standard rate to latest batch rate
                    
                    # Calculate cumulative vendor details
                    auto_update = request.POST.get('auto_update_vendor_due') == 'on'
                    if auto_update:
                        current_outstanding = purchase_total - add_paid
                        product.total_vendor_amount = old_total + purchase_total
                        product.amount_paid_to_vendor = old_amount_paid + add_paid
                        product.remaining_amount_to_vendor = old_remaining + current_outstanding
                    else:
                        product.total_vendor_amount = old_total
                        product.amount_paid_to_vendor = old_amount_paid + add_paid
                        product.remaining_amount_to_vendor = max(0.00, old_remaining - add_paid)
                    
                    product.save()
                    
                    # Log stock history
                    StockHistory.objects.create(
                        product=product,
                        date_entered=add_date,
                        qty_added=add_qty,
                        qty_after=product.stock_quantity,
                    )
                else:
                    # Regular update of other fields without adding a new batch
                    # Determine new paid value submitted
                    new_amount_paid = float(product.amount_paid_to_vendor or 0.00)
                    payment_made = new_amount_paid - old_amount_paid
                    
                    qty_diff = product.initial_quantity - old_initial_qty
                    if qty_diff != 0:
                        product.stock_quantity = max(0, product.stock_quantity + qty_diff)
                    
                    # Prevent overwriting or erasing previous balance and only subtract the payment difference
                    product.total_vendor_amount = old_total
                    product.remaining_amount_to_vendor = max(0.00, old_remaining - payment_made)
                    product.save()
                    
                    if qty_diff != 0:
                        # Find/update the first batch or log simple history
                        first_batch = product.batches.order_by('created_at').first()
                        if first_batch:
                            first_batch.initial_qty = max(0, first_batch.initial_qty + qty_diff)
                            first_batch.current_qty = max(0, first_batch.current_qty + qty_diff)
                            first_batch.save()
                        
                        StockHistory.objects.create(
                            product=product,
                            date_entered=product.stock_entered or timezone.localdate(),
                            qty_added=qty_diff,
                            qty_after=product.stock_quantity,
                        )

            messages.success(request, f"Stock item '{product.name}' was updated successfully.")
            return redirect('stock_add')
    else:
        form = StockForm(instance=product)
        form.initial['stock_entered'] = timezone.localdate()
        
    stock_history = product.stock_history.all()
    batches = product.batches.all().order_by('-created_at')

    # Fetch unique company names with their latest vendor details to support auto-fill
    existing_vendors_data = []
    seen_companies = set()
    for p in Product.objects.exclude(company_name__isnull=True).exclude(company_name="").order_by('-id'):
        cname = p.company_name.strip()
        cname_lower = cname.lower()
        if cname_lower not in seen_companies:
            seen_companies.add(cname_lower)
            existing_vendors_data.append({
                'company_name': cname,
                'due_date': p.vendor_due_date.strftime('%Y-%m-%d') if p.vendor_due_date else '',
                'total_vendor_amount': float(p.total_vendor_amount) if p.total_vendor_amount is not None else '',
                'amount_paid_to_vendor': float(p.amount_paid_to_vendor) if p.amount_paid_to_vendor is not None else '',
                'remaining_amount_to_vendor': float(p.remaining_amount_to_vendor) if p.remaining_amount_to_vendor is not None else '',
            })

    context = {
        'form': form,
        'products': products,
        'categories': categories,
        'product': product,
        'is_edit': True,
        'stock_history': stock_history,
        'batches': batches,
        'existing_vendors_json': json.dumps(existing_vendors_data),
    }
    return render(request, 'stock.html', context)


@admin_required
def stock_delete_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.warning(request, f"Stock item '{name}' was deleted.")
    
    referer = request.META.get('HTTP_REFERER')
    if referer and 'stock1' in referer:
        return redirect('stock1')
    return redirect('stock_add')


@admin_required
def revenue_report_view(request):
    preset = request.GET.get('preset', 'monthly')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.localdate()
    start_date = today
    end_date = today
    
    if preset == 'daily':
        start_date = today
        end_date = today
    elif preset == 'monthly':
        start_date = today.replace(day=1)
        end_date = today
    elif preset == '3months':
        start_date = today - datetime.timedelta(days=90)
        end_date = today
    elif preset == '6months':
        start_date = today - datetime.timedelta(days=180)
        end_date = today
    elif preset == '9months':
        start_date = today - datetime.timedelta(days=270)
        end_date = today
    elif preset == 'yearly':
        start_date = today - datetime.timedelta(days=365)
        end_date = today
    elif preset == 'custom':
        if start_date_str and end_date_str:
            try:
                start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
    else:
        # Default fallback
        preset = 'monthly'
        start_date = today.replace(day=1)
        end_date = today

    # Ensure start_date is not after end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Convert local dates to timezone-aware datetime limits
    start_dt = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    # Query matching bills
    bills = Bill.objects.filter(created_at__range=(start_dt, end_dt)).order_by('-created_at')

    # Calculate metrics
    total_revenue = float(bills.aggregate(total=Sum('grand_total'))['total'] or 0.0)
    total_sales = bills.count()
    total_paid = float(bills.aggregate(total_paid=Sum('amount_given'))['total_paid'] or 0.0)
    avg_order_value = total_revenue / total_sales if total_sales > 0 else 0.0
    total_dues = float(sum(abs(b.amount_to_be_given) for b in bills if b.amount_to_be_given < 0))

    # Group revenue by date for Chart.js trend lines
    chart_labels = []
    chart_values = []
    
    delta = end_date - start_date
    if delta.days <= 45:
        # Group by day
        day_map = {}
        curr = start_date
        while curr <= end_date:
            day_map[curr] = 0.0
            curr += datetime.timedelta(days=1)
            
        for bill in bills:
            local_bill_date = timezone.localtime(bill.created_at).date()
            if local_bill_date in day_map:
                day_map[local_bill_date] += float(bill.grand_total)
                
        # Format labels nicely
        for d in sorted(day_map.keys()):
            chart_labels.append(d.strftime("%b %d"))
            chart_values.append(day_map[d])
    else:
        # Group by month
        month_map = {}
        curr = start_date
        while curr <= end_date:
            m_key = (curr.year, curr.month)
            month_map[m_key] = 0.0
            if curr.month == 12:
                curr = datetime.date(curr.year + 1, 1, 1)
            else:
                curr = datetime.date(curr.year, curr.month + 1, 1)
                
        # Fill data
        for bill in bills:
            local_bill_dt = timezone.localtime(bill.created_at)
            m_key = (local_bill_dt.year, local_bill_dt.month)
            if m_key in month_map:
                month_map[m_key] += float(bill.grand_total)
                
        # Format labels nicely
        for m in sorted(month_map.keys()):
            temp_date = datetime.date(m[0], m[1], 1)
            chart_labels.append(temp_date.strftime("%b %Y"))
            chart_values.append(month_map[m])

    context = {
        'preset': preset,
        'start_date': start_date,
        'end_date': end_date,
        'bills': bills,
        'total_revenue': total_revenue,
        'total_sales': total_sales,
        'total_paid': total_paid,
        'avg_order_value': avg_order_value,
        'total_dues': total_dues,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
    }
    return render(request, 'revenue_report.html', context)


# -------------------------------------------------------------
# DOWNLOAD VIEWS
# -------------------------------------------------------------

def _get_report_bills(request):
    """Shared helper: parse date params and return filtered bills + metadata."""
    preset = request.GET.get('preset', 'monthly')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = timezone.localdate()
    start_date = today
    end_date = today

    if preset == 'daily':
        start_date = end_date = today
    elif preset == 'monthly':
        start_date = today.replace(day=1)
    elif preset == '3months':
        start_date = today - datetime.timedelta(days=90)
    elif preset == '6months':
        start_date = today - datetime.timedelta(days=180)
    elif preset == '9months':
        start_date = today - datetime.timedelta(days=270)
    elif preset == 'yearly':
        start_date = today - datetime.timedelta(days=365)
    elif preset == 'custom' and start_date_str and end_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    else:
        preset = 'monthly'
        start_date = today.replace(day=1)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    start_dt = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))
    bills = Bill.objects.filter(created_at__range=(start_dt, end_dt)).order_by('created_at').prefetch_related('items__product')
    return bills, start_date, end_date


# -- Revenue Report: Excel Download --------------------------
@admin_required
def revenue_report_excel(request):
    bills, start_date, end_date = _get_report_bills(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Report"

    # Styles
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill    = PatternFill("solid", fgColor="EEF2FF")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    right       = Alignment(horizontal="right")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Rukmini Enterprises \u2014 Revenue Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center

    ws.append([])

    # Summary row
    total_revenue = float(bills.aggregate(t=Sum('grand_total'))['t'] or 0)
    total_paid    = float(bills.aggregate(t=Sum('amount_given'))['t'] or 0)
    total_dues    = float(sum(abs(b.amount_to_be_given) for b in bills if b.amount_to_be_given < 0))
    total_sales   = bills.count()

    ws.append(["Total Sales", total_sales, "", "Total Revenue", f"Rs.{total_revenue:.2f}", "", "Amount Collected", f"Rs.{total_paid:.2f}"])
    ws.append(["Amount Balance", f"Rs.{total_dues:.2f}"])
    for cell in ws[4] + ws[5]:
        cell.font = bold
        cell.fill = sub_fill
    ws.append([])

    # Column headers \u2014 one row per bill item
    headers = ["Bill#", "Date", "Customer", "Product", "Qty", "Rate (Rs.)", "Item Total (Rs.)", "Bill Total (Rs.)", "Paid (Rs.)", "Balance (Rs.)", "Status"]
    ws.append(headers)
    for col, cell in enumerate(ws[7], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data rows \u2014 one row per BillItem
    for bill in bills:
        status  = "Paid" if bill.amount_to_be_given >= 0 else "Due"
        balance = float(abs(bill.amount_to_be_given)) if bill.amount_to_be_given < 0 else 0
        items   = list(bill.items.select_related('product').all())
        for idx, item in enumerate(items):
            ws.append([
                bill.id if idx == 0 else "",
                timezone.localtime(bill.created_at).strftime("%d %b %Y %I:%M %p") if idx == 0 else "",
                bill.customer_name if idx == 0 else "",
                item.product.name if item.product else "Deleted",
                item.quantity,
                float(item.rate),
                float(item.total),
                float(bill.grand_total) if idx == 0 else "",
                float(bill.amount_given) if idx == 0 else "",
                balance if idx == 0 else "",
                status if idx == 0 else "",
            ])
            row = ws.max_row
            for cell in ws[row]:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

    # Column widths
    col_widths = [6, 22, 22, 24, 6, 12, 14, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"revenue_report_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -- Revenue Report: PDF Download -----------------------------
@admin_required
def revenue_report_pdf(request):
    bills, start_date, end_date = _get_report_bills(request)

    buffer = io.BytesIO()
    # A4 landscape usable width \u2248 27.7cm after margins
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = []

    title_style = ParagraphStyle('title', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=3)
    sub_style   = ParagraphStyle('sub',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=12)

    elements.append(Paragraph("Rukmini Enterprises - Revenue Report", title_style))
    elements.append(Paragraph(f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}", sub_style))

    # Summary box \u2014 4 label+value pairs
    total_revenue = float(bills.aggregate(t=Sum('grand_total'))['t'] or 0)
    total_paid    = float(bills.aggregate(t=Sum('amount_given'))['t'] or 0)
    total_dues    = float(sum(abs(b.amount_to_be_given) for b in bills if b.amount_to_be_given < 0))
    total_sales   = bills.count()

    summary_data = [
        ["Total Sales", str(total_sales),
         "Total Revenue", f"Rs. {total_revenue:,.2f}",
         "Amount Collected", f"Rs. {total_paid:,.2f}",
         "Amount Balance", f"Rs. {total_dues:,.2f}"],
    ]
    # 8 equal columns across full width
    sw = [3.46*cm] * 8
    summary_table = Table(summary_data, colWidths=sw)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#EEF2FF')),
        ('FONTNAME',      (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8.5),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('BOX',           (0,0), (-1,-1), 0.6, colors.HexColor('#C7D2FE')),
        ('INNERGRID',     (0,0), (-1,-1), 0.4, colors.HexColor('#C7D2FE')),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.35*cm))

    # Detail table
    col_headers = ["Bill#", "Date", "Customer", "Product", "Qty", "Rate", "Item Total", "Bill Total", "Paid", "Balance", "Status"]
    col_widths_pdf = [1.0*cm, 2.6*cm, 3.6*cm, 3.8*cm, 1.2*cm, 2.6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.6*cm]
    data = [col_headers]
    for bill in bills:
        status  = "Paid" if bill.amount_to_be_given >= 0 else "Due"
        balance = f"Rs.{abs(float(bill.amount_to_be_given)):,.2f}" if bill.amount_to_be_given < 0 else "-"
        items   = list(bill.items.select_related('product').all())
        for idx, item in enumerate(items):
            data.append([
                str(bill.id) if idx == 0 else "",
                timezone.localtime(bill.created_at).strftime("%d %b %y") if idx == 0 else "",
                bill.customer_name[:18] if idx == 0 else "",
                (item.product.name[:20] if item.product else "Deleted"),
                str(item.quantity),
                f"Rs.{float(item.rate):,.2f}",
                f"Rs.{float(item.total):,.2f}",
                f"Rs.{float(bill.grand_total):,.2f}" if idx == 0 else "",
                f"Rs.{float(bill.amount_given):,.2f}" if idx == 0 else "",
                balance if idx == 0 else "",
                status if idx == 0 else "",
            ])

    if len(data) == 1:
        data.append(["No records in this period."] + [""]*10)

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    tbl.setStyle(TableStyle([
        # -- Header row --
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  8),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
        ('VALIGN',        (0,0), (-1,0),  'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  6),
        ('BOTTOMPADDING', (0,0), (-1,0),  6),
        # -- Data rows --
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 7.5),
        ('VALIGN',        (0,1), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        # Left-align text columns: #, Date, Customer, Product
        ('ALIGN',         (0,1), (0,-1),  'CENTER'),   # #
        ('ALIGN',         (1,1), (1,-1),  'LEFT'),     # Date
        ('ALIGN',         (2,1), (2,-1),  'LEFT'),     # Customer
        ('ALIGN',         (3,1), (3,-1),  'LEFT'),     # Product
        # Center Qty
        ('ALIGN',         (4,1), (4,-1),  'CENTER'),   # Qty
        # Right-align numeric columns
        ('ALIGN',         (5,1), (8,-1),  'RIGHT'),    # Rate, Total, Paid, Balance
        # Center Status
        ('ALIGN',         (9,1), (9,-1),  'CENTER'),   # Status
        # Alternating row colors
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        # Borders
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)
    filename = f"revenue_report_{start_date}_{end_date}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# -- Stock Report: Excel Download -----------------------------
@admin_required
def stock_report_excel(request):
    products = Product.objects.all().select_related('category').order_by('category__name', 'name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    low_fill    = PatternFill("solid", fgColor="FEE2E2")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    batch_font  = Font(italic=True, size=10, color="475569")
    batch_fill  = PatternFill("solid", fgColor="F8FAFC")

    ws.merge_cells("A1:O1")
    ws["A1"] = "Rukmini Enterprises \u2014 Stock Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Generated: {timezone.localdate().strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center
    ws.append([])

    headers = [
        "#", "Product Name", "Category", "Unit / Size", "Company", "Stock Entered", 
        "Unit Cost (\u20b9)", "Total Vendor Amt (\u20b9)", "Paid to Vendor (\u20b9)", "Remaining Due (\u20b9)", 
        "Due Date", "Selling Price (\u20b9)", "Initial Qty", "Stock Qty", "Status"
    ]
    ws.append(headers)
    for col, cell in enumerate(ws[4], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    idx = 1
    for p in products:
        status = "Low Stock" if p.is_low_stock else "OK"
        ws.append([
            idx,
            p.name,
            p.category.name if p.category else "\u2014",
            p.size or "\u2014",
            p.company_name or "\u2014",
            p.stock_entered.strftime('%d %b %Y') if p.stock_entered else "\u2014",
            float(p.vendor_cost or 0.00),
            float(p.total_vendor_amount or 0.00),
            float(p.amount_paid_to_vendor or 0.00),
            float(p.remaining_amount_to_vendor or 0.00),
            p.vendor_due_date.strftime('%d %b %Y') if p.vendor_due_date else "\u2014",
            float(p.selling_price or 0.00),
            p.initial_quantity,
            p.stock_quantity,
            status,
        ])
        idx += 1
        row = ws.max_row
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if p.is_low_stock:
            for cell in ws[row]:
                cell.fill = low_fill

        # Add active batches
        active_batches = p.batches.filter(current_qty__gt=0).order_by('created_at')
        for b in active_batches:
            ws.append([
                "",
                f"  + Batch (Recd: {b.stock_entered.strftime('%d %b %Y') if b.stock_entered else '\u2014'})",
                "",
                "",
                "",
                "",
                float(b.purchase_rate or 0.00),
                "",
                "",
                "",
                "",
                float(b.selling_price or 0.00),
                b.initial_qty,
                b.current_qty,
                "",
            ])
            brow = ws.max_row
            for cell in ws[brow]:
                cell.border = border
                cell.font = batch_font
                cell.fill = batch_fill
                cell.alignment = Alignment(vertical="center")

    col_widths = [5, 30, 18, 12, 22, 15, 15, 18, 18, 18, 15, 16, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="stock_report_{timezone.localdate()}.xlsx"'
    wb.save(response)
    return response


@admin_required
def stock_report_pdf(request):
    products = Product.objects.all().select_related('category').order_by('category__name', 'name')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.1*cm, leftMargin=1.1*cm,
                            topMargin=1.1*cm, bottomMargin=1.1*cm)

    elements = []

    title_style = ParagraphStyle('title2', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#0F172A'), alignment=TA_CENTER, spaceAfter=3)
    sub_style   = ParagraphStyle('sub2',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=12)

    elements.append(Paragraph("Rukmini Enterprises - Stock Report", title_style))
    elements.append(Paragraph(f"Generated: {timezone.localdate().strftime('%d %b %Y')}", sub_style))

    # Custom typography style guidelines to ensure cell text wrap and alignments
    styles = getSampleStyleSheet()
    
    hdr_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.5,
        textColor=colors.white,
        alignment=TA_CENTER
    )
    cell_l = ParagraphStyle(
        'CellLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.5,
        textColor=colors.HexColor('#1E293B'),
        alignment=TA_LEFT
    )
    cell_c = ParagraphStyle(
        'CellCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.5,
        textColor=colors.HexColor('#1E293B'),
        alignment=TA_CENTER
    )
    cell_r = ParagraphStyle(
        'CellRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.5,
        textColor=colors.HexColor('#1E293B'),
        alignment=TA_RIGHT
    )
    cell_batch = ParagraphStyle(
        'CellBatch',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=6,
        textColor=colors.HexColor('#475569'),
        alignment=TA_LEFT
    )
    cell_batch_r = ParagraphStyle(
        'CellBatchRight',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=6,
        textColor=colors.HexColor('#475569'),
        alignment=TA_RIGHT
    )
    cell_batch_c = ParagraphStyle(
        'CellBatchCenter',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=6,
        textColor=colors.HexColor('#475569'),
        alignment=TA_CENTER
    )
    status_ok = ParagraphStyle(
        'StatusOk',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.5,
        textColor=colors.HexColor('#059669'),
        alignment=TA_CENTER
    )
    status_low = ParagraphStyle(
        'StatusLow',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.5,
        textColor=colors.HexColor('#DC2626'),
        alignment=TA_CENTER
    )

    # 15 Columns. Total width = 27.5 cm (fits landscape A4 printable area of 27.7cm)
    col_widths_pdf = [0.8*cm, 3.5*cm, 2.0*cm, 1.3*cm, 2.2*cm, 2.0*cm, 1.8*cm, 2.0*cm, 1.8*cm, 1.8*cm, 2.0*cm, 1.8*cm, 1.5*cm, 1.5*cm, 1.5*cm]

    col_headers = [
        Paragraph("<b>#</b>", hdr_style),
        Paragraph("<b>Product Name</b>", hdr_style),
        Paragraph("<b>Category</b>", hdr_style),
        Paragraph("<b>Size</b>", hdr_style),
        Paragraph("<b>Company</b>", hdr_style),
        Paragraph("<b>Entered</b>", hdr_style),
        Paragraph("<b>Unit Cost</b>", hdr_style),
        Paragraph("<b>Total Amt</b>", hdr_style),
        Paragraph("<b>Paid</b>", hdr_style),
        Paragraph("<b>Due</b>", hdr_style),
        Paragraph("<b>Due Date</b>", hdr_style),
        Paragraph("<b>Selling</b>", hdr_style),
        Paragraph("<b>Initial</b>", hdr_style),
        Paragraph("<b>Stock</b>", hdr_style),
        Paragraph("<b>Status</b>", hdr_style),
    ]
    data = [col_headers]
    
    t_styles = [
        # -- Header row styling --
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#1E293B')), # Professional dark slate header
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('VALIGN',        (0,0), (-1,0),  'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  6),
        ('BOTTOMPADDING', (0,0), (-1,0),  6),
    ]

    idx = 1
    for p in products:
        status_lbl = "Low Stock" if p.is_low_stock else "OK"
        p_status = Paragraph("OK", status_ok) if status_lbl == "OK" else Paragraph("Low Stock", status_low)
        
        row_idx = len(data)
        data.append([
            Paragraph(str(idx), cell_c),
            Paragraph(p.name, cell_l),
            Paragraph(p.category.name if p.category else "-", cell_l),
            Paragraph(p.size or "-", cell_c),
            Paragraph(p.company_name or "-", cell_l),
            Paragraph(p.stock_entered.strftime('%d %b %Y') if p.stock_entered else "-", cell_c),
            Paragraph(f"Rs.{float(p.vendor_cost or 0.00):,.2f}", cell_r),
            Paragraph(f"Rs.{float(p.total_vendor_amount or 0.00):,.2f}", cell_r),
            Paragraph(f"Rs.{float(p.amount_paid_to_vendor or 0.00):,.2f}", cell_r),
            Paragraph(f"Rs.{float(p.remaining_amount_to_vendor or 0.00):,.2f}", cell_r),
            Paragraph(p.vendor_due_date.strftime('%d %b %Y') if p.vendor_due_date else "-", cell_c),
            Paragraph(f"Rs.{float(p.selling_price or 0.00):,.2f}", cell_r),
            Paragraph(str(p.initial_quantity), cell_c),
            Paragraph(str(p.stock_quantity), cell_c),
            p_status,
        ])
        
        # Style for product row
        if p.is_low_stock:
            t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#FEF2F2'))) # Soft red warning background
        else:
            bg_color = colors.white if idx % 2 != 0 else colors.HexColor('#F8FAFC')
            t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))
            
        t_styles.extend([
            ('VALIGN',        (0, row_idx), (-1, row_idx), 'MIDDLE'),
            ('TOPPADDING',    (0, row_idx), (-1, row_idx), 4),
            ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 4),
        ])
        
        idx += 1
        
        # Active Batches
        active_batches = p.batches.filter(current_qty__gt=0).order_by('created_at')
        for b in active_batches:
            batch_row_idx = len(data)
            dt_str = b.stock_entered.strftime('%d %b %Y') if b.stock_entered else '-'
            data.append([
                Paragraph("", cell_batch_c),
                Paragraph(f"  \u21b3 Batch (Recd: {dt_str})", cell_batch),
                Paragraph("", cell_batch),
                Paragraph("", cell_batch_c),
                Paragraph("", cell_batch),
                Paragraph("", cell_batch_c),
                Paragraph(f"Rs.{float(b.purchase_rate or 0.00):,.2f}", cell_batch_r),
                Paragraph("", cell_batch_r),
                Paragraph("", cell_batch_r),
                Paragraph("", cell_batch_r),
                Paragraph("", cell_batch_c),
                Paragraph(f"Rs.{float(b.selling_price or 0.00):,.2f}", cell_batch_r),
                Paragraph(str(b.initial_qty), cell_batch_c),
                Paragraph(str(b.current_qty), cell_batch_c),
                Paragraph("", cell_batch_c),
            ])
            t_styles.extend([
                ('BACKGROUND',    (0, batch_row_idx), (-1, batch_row_idx), colors.HexColor('#F8FAFC')),
                ('VALIGN',        (0, batch_row_idx), (-1, batch_row_idx), 'MIDDLE'),
                ('TOPPADDING',    (0, batch_row_idx), (-1, batch_row_idx), 3.5),
                ('BOTTOMPADDING', (0, batch_row_idx), (-1, batch_row_idx), 3.5),
            ])

    if len(data) == 1:
        data.append([Paragraph("No stock items found.", cell_c)] + [""] * 14)
        t_styles.append(('SPAN', (0, 1), (-1, 1)))

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    
    # Subtle borders and compact padding
    t_styles.extend([
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
    ])
    
    tbl.setStyle(TableStyle(t_styles))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_report_{timezone.localdate()}.pdf"'
    return response


# -- Stock History: Excel Download ----------------------------
@admin_required
def stock_history_excel(request, pk):
    product = get_object_or_404(Product, pk=pk)
    history = product.stock_history.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock History"

    # Styles
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    add_fill    = PatternFill("solid", fgColor="D1FAE5")   # green tint for additions
    rem_fill    = PatternFill("solid", fgColor="FEE2E2")   # red tint for removals
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Rukmini Enterprises \u2014 Stock Update History"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Product: {product.name}  |  Company: {product.company_name or '\u2014'}  |  Category: {product.category.name if product.category else '\u2014'}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center

    ws.merge_cells("A3:E3")
    ws["A3"] = f"Generated: {timezone.localdate().strftime('%d %b %Y')}  |  Current Stock: {product.stock_quantity} units"
    ws["A3"].font = Font(size=10, color="64748B")
    ws["A3"].alignment = center
    ws.append([])

    # Column headers
    headers = ["#", "Date of Stock Entry", "Qty Added / Removed", "Total Stock After Update", "Logged At"]
    ws.append(headers)
    for col, cell in enumerate(ws[5], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data rows
    for i, entry in enumerate(history, 1):
        ws.append([
            i,
            entry.date_entered.strftime("%d %b %Y"),
            f"+{entry.qty_added}" if entry.qty_added > 0 else str(entry.qty_added),
            entry.qty_after,
            timezone.localtime(entry.recorded_at).strftime("%d %b %Y, %I:%M %p"),
        ])
        row = ws.max_row
        fill = add_fill if entry.qty_added > 0 else rem_fill
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center")
        # Colour the qty column
        ws.cell(row=row, column=3).fill = fill

    if not history.exists():
        ws.append(["", "No history records found.", "", "", ""])

    # Column widths
    for i, w in enumerate([5, 22, 22, 26, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    safe_name = product.name.replace(" ", "_")[:30]
    response["Content-Disposition"] = f'attachment; filename="stock_history_{safe_name}_{timezone.localdate()}.xlsx"'
    wb.save(response)
    return response


# -- Stock History: PDF Download ----------------------------
@admin_required
def stock_history_pdf(request, pk):
    product = get_object_or_404(Product, pk=pk)
    history = product.stock_history.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []

    title_style = ParagraphStyle('htitle', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('hsub',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=4)

    elements.append(Paragraph("Rukmini Enterprises - Stock Update History", title_style))
    elements.append(Paragraph(
        f"Product: {product.name}  |  Company: {product.company_name or '\u2014'}  |  Category: {product.category.name if product.category else '\u2014'}",
        sub_style
    ))
    elements.append(Paragraph(
        f"Generated: {timezone.localdate().strftime('%d %b %Y')}   |   Current Stock: {product.stock_quantity} units",
        sub_style
    ))
    elements.append(Spacer(1, 0.4*cm))

    # Summary strip
    total_added   = sum(e.qty_added for e in history if e.qty_added > 0)
    total_removed = abs(sum(e.qty_added for e in history if e.qty_added < 0))
    summary_data = [[
        "Total Updates", str(history.count()),
        "Total Added", str(total_added),
        "Total Removed", str(total_removed),
        "Current Stock", str(product.stock_quantity),
    ]]
    sw = [3.46*cm] * 8
    summary_tbl = Table(summary_data, colWidths=sw)
    summary_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#EEF2FF')),
        ('FONTNAME',      (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8.5),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('BOX',           (0,0), (-1,-1), 0.6, colors.HexColor('#C7D2FE')),
        ('INNERGRID',     (0,0), (-1,-1), 0.4, colors.HexColor('#C7D2FE')),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1, 0.4*cm))

    # Detail table
    col_headers = ["#", "Date of Stock Entry", "Qty Added / Removed", "Total Stock After Update", "Logged At"]
    col_widths_pdf = [1.2*cm, 5.5*cm, 5.5*cm, 6.0*cm, 6.5*cm]
    data = [col_headers]

    for i, entry in enumerate(history, 1):
        qty_text = f"+{entry.qty_added}" if entry.qty_added > 0 else str(entry.qty_added)
        data.append([
            str(i),
            entry.date_entered.strftime("%d %b %Y"),
            qty_text,
            f"{entry.qty_after} units",
            timezone.localtime(entry.recorded_at).strftime("%d %b %Y, %I:%M %p"),
        ])

    if len(data) == 1:
        data.append(["", "No history records found.", "", "", ""])

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)

    # Build row-by-row colours for added vs removed
    row_styles = [
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  9),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  7),
        ('BOTTOMPADDING', (0,0), (-1,0),  7),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 8.5),
        ('TOPPADDING',    (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ('ALIGN',         (0,1), (0,-1),  'CENTER'),
        ('ALIGN',         (1,1), (1,-1),  'CENTER'),
        ('ALIGN',         (2,1), (2,-1),  'CENTER'),
        ('ALIGN',         (3,1), (3,-1),  'CENTER'),
        ('ALIGN',         (4,1), (4,-1),  'CENTER'),
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING',   (0,0), (-1,-1), 5),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
    ]
    # Colour qty column per row
    for i, entry in enumerate(history, 1):
        if entry.qty_added > 0:
            row_styles.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#065F46')))
            row_styles.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#D1FAE5')))
        else:
            row_styles.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#991B1B')))
            row_styles.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FEE2E2')))
        # Alternating row bg
        if i % 2 == 0:
            row_styles.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#F8FAFC')))
            row_styles.append(('BACKGROUND', (3, i), (4, i), colors.HexColor('#F8FAFC')))

    tbl.setStyle(TableStyle(row_styles))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)
    safe_name = product.name.replace(" ", "_")[:30]
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_history_{safe_name}_{timezone.localdate()}.pdf"'
    return response


# -- All Stock History: Excel Download ----------------------------
@admin_required
def stock_all_history_excel(request):
    from .models import StockHistory
    history_qs = StockHistory.objects.select_related('product').order_by('product__name', '-recorded_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock History"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    group_fill  = PatternFill("solid", fgColor="EEF2FF")
    add_fill    = PatternFill("solid", fgColor="D1FAE5")
    rem_fill    = PatternFill("solid", fgColor="FEE2E2")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Rukmini Enterprises \u2014 All Stock Update History"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {timezone.localdate().strftime('%d %b %Y')}  |  Total Records: {history_qs.count()}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = center
    ws.append([])

    # Column headers
    headers = ["#", "Product Name", "Date of Stock Entry", "Qty Added / Removed", "Total Stock After", "Logged At"]
    ws.append(headers)
    for col, cell in enumerate(ws[4], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    current_product = None
    row_num = 0
    for entry in history_qs:
        # Group header row per product
        if entry.product != current_product:
            current_product = entry.product
            ws.append([
                "", f"\u25b6  {entry.product.name}  \u2014  {entry.product.company_name or '\u2014'}",
                "", "", "", ""
            ])
            grp_row = ws.max_row
            ws.merge_cells(f"B{grp_row}:F{grp_row}")
            for cell in ws[grp_row]:
                cell.fill = group_fill
                cell.font = Font(bold=True, size=10, color="4F46E5")
                cell.alignment = Alignment(vertical="center")

        row_num += 1
        qty_str = f"+{entry.qty_added}" if entry.qty_added > 0 else str(entry.qty_added)
        ws.append([
            row_num,
            entry.product.name,
            entry.date_entered.strftime("%d %b %Y"),
            qty_str,
            entry.qty_after,
            timezone.localtime(entry.recorded_at).strftime("%d %b %Y, %I:%M %p"),
        ])
        row = ws.max_row
        fill = add_fill if entry.qty_added > 0 else rem_fill
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=row, column=4).fill = fill

    if not history_qs.exists():
        ws.append(["", "No stock history records found.", "", "", "", ""])

    for i, w in enumerate([5, 28, 22, 22, 22, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="all_stock_history_{timezone.localdate()}.xlsx"'
    wb.save(response)
    return response


# -- All Stock History: PDF Download ----------------------------
@admin_required
def stock_all_history_pdf(request):
    from .models import StockHistory
    history_qs = StockHistory.objects.select_related('product').order_by('product__name', '-recorded_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []

    title_style = ParagraphStyle('ahtitle', fontSize=15, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('ahsub', fontSize=9, fontName='Helvetica',
                                  textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=12)
    group_style = ParagraphStyle('ahgrp', fontSize=9, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#4F46E5'))

    elements.append(Paragraph("Rukmini Enterprises - All Stock Update History", title_style))
    elements.append(Paragraph(
        f"Generated: {timezone.localdate().strftime('%d %b %Y')}   |   Total Records: {history_qs.count()}",
        sub_style
    ))

    # Summary strip
    total_added   = sum(e.qty_added for e in history_qs if e.qty_added > 0)
    total_removed = abs(sum(e.qty_added for e in history_qs if e.qty_added < 0))
    summary_data = [[
        "Total Records", str(history_qs.count()),
        "Total Added", str(total_added),
        "Total Removed", str(total_removed),
        "Products Tracked", str(history_qs.values('product').distinct().count()),
    ]]
    sw = [3.46*cm] * 8
    summary_tbl = Table(summary_data, colWidths=sw)
    summary_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#EEF2FF')),
        ('FONTNAME',      (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8.5),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('BOX',           (0,0), (-1,-1), 0.6, colors.HexColor('#C7D2FE')),
        ('INNERGRID',     (0,0), (-1,-1), 0.4, colors.HexColor('#C7D2FE')),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1, 0.4*cm))

    # Detail table
    col_headers = ["#", "Product Name", "Date of Entry", "Qty Added / Removed", "Stock After", "Logged At"]
    col_widths_pdf = [1.0*cm, 6.5*cm, 4.5*cm, 4.5*cm, 3.5*cm, 5.7*cm]
    data = [col_headers]
    row_styles = [
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  9),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  6),
        ('BOTTOMPADDING', (0,0), (-1,0),  6),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('TOPPADDING',    (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('ALIGN',         (0,1), (0,-1),  'CENTER'),
        ('ALIGN',         (1,1), (1,-1),  'LEFT'),
        ('ALIGN',         (2,1), (4,-1),  'CENTER'),
        ('ALIGN',         (5,1), (5,-1),  'CENTER'),
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
    ]

    current_product = None
    row_num = 0
    data_row_index = 1  # header is index 0

    for entry in history_qs:
        # Group header row per product
        if entry.product != current_product:
            current_product = entry.product
            data.append([f"\u25b6  {entry.product.name}  \u2014  {entry.product.company_name or '\u2014'}", "", "", "", "", ""])
            row_styles.append(('BACKGROUND', (0, data_row_index), (-1, data_row_index), colors.HexColor('#EEF2FF')))
            row_styles.append(('TEXTCOLOR',  (0, data_row_index), (-1, data_row_index), colors.HexColor('#4F46E5')))
            row_styles.append(('FONTNAME',   (0, data_row_index), (-1, data_row_index), 'Helvetica-Bold'))
            row_styles.append(('SPAN',       (0, data_row_index), (-1, data_row_index)))
            data_row_index += 1

        row_num += 1
        qty_str = f"+{entry.qty_added}" if entry.qty_added > 0 else str(entry.qty_added)
        data.append([
            str(row_num),
            entry.product.name[:28],
            entry.date_entered.strftime("%d %b %Y"),
            qty_str,
            f"{entry.qty_after} units",
            timezone.localtime(entry.recorded_at).strftime("%d %b %y %I:%M %p"),
        ])
        if entry.qty_added > 0:
            row_styles.append(('TEXTCOLOR',  (3, data_row_index), (3, data_row_index), colors.HexColor('#065F46')))
            row_styles.append(('BACKGROUND', (3, data_row_index), (3, data_row_index), colors.HexColor('#D1FAE5')))
        else:
            row_styles.append(('TEXTCOLOR',  (3, data_row_index), (3, data_row_index), colors.HexColor('#991B1B')))
            row_styles.append(('BACKGROUND', (3, data_row_index), (3, data_row_index), colors.HexColor('#FEE2E2')))
        if row_num % 2 == 0:
            row_styles.append(('BACKGROUND', (0, data_row_index), (2, data_row_index), colors.HexColor('#F8FAFC')))
            row_styles.append(('BACKGROUND', (4, data_row_index), (5, data_row_index), colors.HexColor('#F8FAFC')))
        data_row_index += 1

    if len(data) == 1:
        data.append(["", "No stock history records found.", "", "", "", ""])

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    tbl.setStyle(TableStyle(row_styles))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="all_stock_history_{timezone.localdate()}.pdf"'
    return response


@admin_required
def stock_ledger_view(request):
    # Fetch all batches (purchases)
    batches = ProductBatch.objects.select_related('product').all()
    # Fetch all batch sales (sales)
    sales = BatchSale.objects.select_related('batch__product', 'bill_item__bill').all()
    
    transactions = []
    
    # Add purchases to the list
    for b in batches:
        transactions.append({
            'date': b.stock_entered,
            'time': b.created_at,
            'product': b.product.name,
            'type': 'Purchase',
            'qty_in': b.initial_qty,
            'qty_out': 0,
            'rate': b.purchase_rate,
            'remaining_qty': b.current_qty,
            'ref': f"Batch #{b.id}",
        })
        
    # Add sales to the list
    for s in sales:
        transactions.append({
            'date': s.created_at.date(),
            'time': s.created_at,
            'product': s.batch.product.name,
            'type': 'Sale',
            'qty_in': 0,
            'qty_out': s.quantity_sold,
            'rate': s.purchase_rate,
            'remaining_qty': s.batch.current_qty,
            'ref': f"Bill #{s.bill_item.bill.id}",
        })
        
    # Sort transactions by time descending (newest first)
    transactions.sort(key=lambda x: x['time'], reverse=True)
    
    context = {
        'transactions': transactions,
    }
    return render(request, 'stock_ledger.html', context)

